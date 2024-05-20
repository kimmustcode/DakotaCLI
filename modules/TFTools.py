import openpyxl
import json
import os

champs = []

traits = ['Altruist', 'Arcanist', 'Behemoth', 'Bruiser', 'Duelist', 'Exalted', 'Invoker', 'Reaper', 'Sage', 'Sniper', 'Trickshot', 'Warden', 'Dragonlord', 'Dryad', 'Fated', 'Fortune', 'Ghostly', 'Heavenly', 'Inkshadow', 'Mythic', 'Porcelain', 'Storyweaver', 'Umbral']
cat    = r"""
                 ___
             _.-|   |          |\__/,|   (`\
             {  |   |          |o o  |__ _) )
             "-.|___|        _.( T   )  `  /
              .--'-`-.     _((_ `^--' /_<  \
            .+|______|__.-||__)`-'(((/  (((/
        
        """


class Champ:
    def __init__(self, name, traits, cost):
        self.name = name
        self.traits = traits
        self.cost = cost


# Load data from sheet to champ class and shit 
def load_data():
    dataframe = openpyxl.load_workbook("rsc/champs.xlsx")

    dataframe1 = dataframe.active 

    # 3-pronged switch 0 - name, 1 - traits, 2 - cost 
    titch = 0 

    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            if titch == 0:
                name = col[row].value
                titch += 1
            elif titch == 1: 
                titch += 1 
                temp = []
                tempTraits = json.loads(col[row].value)
                for trait in tempTraits:
                    temp.append(traits[int(trait)])
            else: 
                titch = 0
                cost = col[row].value
                champs.append(Champ(name, temp, cost))

    count = 1 

    # Bubble sort (lol) sorts champs in ascending order by cost
    while count > 0: 
        count = 0 
        for i in range(len(champs) - 1):
            if champs[i].cost > champs[i + 1].cost:
                temp = champs[i]
                champs[i] = champs[i + 1]
                champs[i + 1] = temp
                count += 1

    return 

# Where the module starts 
def mod_start():
    load_data()
    
    exit1 = False 

    while exit1 == False:
        os.system('cls')
        print(cat)    
        print("What TFT tool do you want?\n--(team) calculator\n--(print) champs\n--(return)") 
        temp = input("\n>")

        if temp == 'team':
            emblem_calc()
        elif temp == 'print':
            print_champs()
        elif temp == 'return':
            exit1 = True 

    return 

# Print champs matching given trait
def print_champs():
    os.system('cls')
    print(cat)    

    # User Input 
    print("What Trait do you want to see?")
    trait = input("\n>")
    print("")

    # Normalize input (make first letter capital if not)
    temp = list(trait)
    if temp[0].islower():
        temp[0] = temp[0].upper()
    trait = ''.join(temp)
    
    # Print characters matching the trait
    if trait == 'All':
        for chara in champs:
            print(chara.name, chara.traits, chara.cost)
    else:
        for chara in champs: 
            if trait in chara.traits:
                print(chara.name, chara.traits, chara.cost)

    input("\nEnter to continue.")

    return 

# Takes up to 3 emblems and gives you the 9 champs that gives you the most completed augments
def emblem_calc():
    os.system('cls')
    print(cat)   
    print("Takes max 3 traits as input to output a team.\nEnter in order of preference.\n")
    testTraits = []

    x = 1 

    # Take in and normalize the traits input
    while x != '0' and len(testTraits) < 3:
        x = input("Enter Trait(Max 3, '0' to exit): ")
        
        temp = list(x)
        if temp[0].islower():
            temp[0] = temp[0].upper()
        x = ''.join(temp)
        testTraits.append(x)

    # Load number of priorities based on how many traits you entered
    priors = [1] * len(testTraits)  # priority 
    team = []

    # First try to find champs with two match traits 
    for champ in champs: 
        cnt = 0 
        for trait in champ.traits:
            if trait in testTraits:
                cnt += 1 
        if cnt >= 2:
            team.append(champ)
        
    # Add new traits to team and update priority 
    for champ in team: 
        testTraits, priors = updatePrior(testTraits, priors, champ)

    # Starting at the highest priority trait find champs matching that trait
    # (testTraits gets updated to reflect the updated prior values each pass)
    for trait in testTraits: 
            for champ in champs:
                if len(team) < 10:
                    if trait in champ.traits and champ not in team:
                        team.append(champ)
                        updatePrior(testTraits, priors, champ)


    count = 1 

    # Same sort to sort team by cost 
    while count > 0: 
        count = 0 
        for i in range(len(team) - 1):
            if team[i].cost > team[i + 1].cost:
                temp = team[i]
                team[i] = team[i + 1]
                team[i + 1] = temp
                count += 1
    

    # Print Team
    for champ in team:
        print(champ.name, champ.traits)
        print("Cost: " + str(champ.cost))
        print("")

    print(testTraits)
    print(priors)
                   
    input("Enter to continue.")

    return


# Takes traits, priors, and the new champ in order to update the team traits and priors 
def updatePrior(traits, priors, champ):

    # If trait isnt in traits add it, otherwise +1 to the trait prior
    for trait in champ.traits:
        if trait in traits:
            priors[traits.index(trait)] += 1
        elif trait not in traits: 
            traits.append(trait)
            priors.append(1)

    cnt = 1

    # Sort in desc order by prior
    while cnt > 0:
        cnt = 0 
        for i in range(len(traits) - 1):
            if priors[i] < priors[i + 1]:
                temp = priors[i]
                priors[i] = priors[i + 1]
                priors[i + 1] = temp
                temp = traits[i]
                traits[i] = traits[i + 1]
                traits[i + 1] = temp
                
                cnt += 1
    return traits, priors